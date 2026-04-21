import streamlit as st
import pandas as pd
import os
from datetime import date

# ─────────────────────────────────────────────
# 1. PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="B&B · Project Financial Suite",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# 2. CUSTOM THEME  (Navy #153275 · Red #D10D38)
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Montserrat:wght@600;700;800&display=swap');
:root {
    --navy:  #153275;
    --red:   #D10D38;
    --slate: #1e293b;
    --muted: #64748b;
    --light: #f1f5f9;
    --card:  #ffffff;
    --border:#e2e8f0;
    --green: #16a34a;
    --amber: #d97706;
}
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
h1, h2, h3, .page-header   { font-family: 'Montserrat', sans-serif !important; }
#MainMenu, footer, header   { visibility: hidden; }
div[data-testid="stDecoration"] { display: none; }
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, var(--navy) 0%, #0f2355 100%);
}
section[data-testid="stSidebar"] * { color: #ffffff !important; }
section[data-testid="stSidebar"] .stRadio label {
    padding: 0.55rem 0.75rem; border-radius: 8px;
    transition: background 0.2s; font-weight: 500; font-size: 0.9rem;
}
section[data-testid="stSidebar"] .stRadio label:hover { background: rgba(255,255,255,0.1); }
section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.15) !important; }
.stButton > button, .stFormSubmitButton > button {
    background: var(--navy) !important; color: #fff !important;
    border: none !important; border-radius: 8px !important;
    padding: 0.5rem 1.25rem !important; font-weight: 600 !important;
    font-size: 0.875rem !important; transition: all 0.2s !important;
}
.stButton > button:hover, .stFormSubmitButton > button:hover {
    background: #1a3f8f !important; transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(21,50,117,0.3) !important;
}
.stTextInput input, .stSelectbox > div > div,
.stDateInput input, .stTextArea textarea {
    border-radius: 8px !important; border: 1.5px solid var(--border) !important;
    font-size: 0.9rem !important;
}
.stDataFrame { border: 1px solid var(--border); border-radius: 10px; overflow: hidden; }
.metric-row  { display: flex; gap: 1rem; margin-bottom: 1.25rem; }
.metric-card {
    flex: 1; background: var(--card); border: 1px solid var(--border);
    border-radius: 12px; padding: 1.1rem 1.25rem; text-align: center;
    transition: box-shadow 0.2s;
}
.metric-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.06); }
.metric-card .label {
    font-size: 0.75rem; font-weight: 600; color: var(--muted);
    text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 0.3rem;
}
.metric-card .value {
    font-family: 'Montserrat', sans-serif; font-size: 1.35rem;
    font-weight: 700; color: var(--slate);
}
.metric-card.navy  .value { color: var(--navy); }
.metric-card.red   .value { color: var(--red);  }
.metric-card.green .value { color: var(--green);}
.status-badge {
    display: inline-block; padding: 0.2rem 0.7rem; border-radius: 20px;
    font-size: 0.78rem; font-weight: 600; letter-spacing: 0.02em;
}
.status-active  { background: #dcfce7; color: #166534; }
.status-pending { background: #fef9c3; color: #854d0e; }
.status-awarded { background: #dbeafe; color: #1e40af; }
.status-closed  { background: #f1f5f9; color: #475569; }
.page-header {
    display: flex; align-items: center; gap: 0.75rem;
    padding-bottom: 0.5rem; margin-bottom: 1.25rem;
    border-bottom: 3px solid var(--navy);
}
.page-header .icon  { font-size: 1.6rem; }
.page-header .title {
    font-family: 'Montserrat', sans-serif; font-size: 1.65rem;
    font-weight: 700; color: var(--slate); margin: 0;
}
.page-header .subtitle { font-size: 0.85rem; color: var(--muted); margin-left: auto; }
.section-divider { height: 1px; background: var(--border); margin: 1.75rem 0; }
.form-card {
    background: var(--card); border: 1px solid var(--border);
    border-radius: 14px; padding: 1.5rem; margin-bottom: 1.25rem;
}
.stSuccess, .stError, .stWarning { border-radius: 10px !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def page_header(icon: str, title: str, subtitle: str = ""):
    sub = f'<span class="subtitle">{subtitle}</span>' if subtitle else ""
    st.markdown(
        f'<div class="page-header">'
        f'  <span class="icon">{icon}</span>'
        f'  <span class="title">{title}</span>{sub}'
        f'</div>',
        unsafe_allow_html=True,
    )

def metric_cards(items: list[tuple]):
    cards = ""
    for label, value, cls in items:
        cards += (
            f'<div class="metric-card {cls}">'
            f'  <div class="label">{label}</div>'
            f'  <div class="value">{value}</div>'
            f'</div>'
        )
    st.markdown(f'<div class="metric-row">{cards}</div>', unsafe_allow_html=True)

def status_badge(status: str) -> str:
    s = str(status).strip().lower()
    cls_map = {"active": "status-active", "pending": "status-pending",
               "awarded": "status-awarded", "closed": "status-closed"}
    cls = cls_map.get(s, "status-pending")
    return f'<span class="status-badge {cls}">{status}</span>'


# ─────────────────────────────────────────────
# FILE CONFIG
# ─────────────────────────────────────────────
MASTER_FILE = "Master_Project_Registry.xlsx"
WIP_FILE    = "WIP.xlsx"

P1_COLS  = ['Status', 'Job #', 'Job Name', 'Client Name']
P2_COLS  = ['Contract Date', 'As Sold Contract Amount', 'Approved Change Orders',
            "CO's Added during WIP", 'Current Contract Amount', 'Retainage %', 'Notes']
P3_COLS  = ['Paula Billings', 'JVM Billings', 'JTD Billings', 'Retainage', 'Net Billings']
ALL_COLS = P1_COLS + P2_COLS + P3_COLS

WIP_TARGET_COLS = [
    'Status', 'Job #', 'Job Name', 'Client Name',
    'As Sold Contract Amount', 'Approved Change Orders', "CO's Added during WIP",
    'Current Contract Amount', 'Retainage %',
    'JTD Billings', 'Retaininage', 'Net Billing',
    '(Over) / Under Billings', 'Over / (Under) Billings %',
    'Total Remaning Billing', 'Billings %',
    'Total Cost Projection (As of report date)',
    'Remaining Cost to Complete Projection (c)',
    'JTD Costs (B)', '% Complete',
    'Marin %\nwriteup/wiritedown', 'Projected Margin %',
    'Write Up / Write Down Potential ', 'Write Up / Write Down Amount',
    'Budgeted Gross Proft ', 'Projected Gross Profit',
    'GP Variance(%)', 'GP Variance (USD)',
    'JTD Revenue', 'Unearned Rev',
    'Remaining Profit Margin',
    'JTD Cost', 'JTD Profits',
    'YTD Revenue', 'YTD Cost', 'YTD Profit',
    'MTD Revenue', 'MTD Cost', 'MTD Profit',
    'YTD Cost.1', 'YTD Revenue.1',
]

COL_RENAME_TO_WIP = {
    'Retainage'             : 'Retaininage',
    'Net Billings'          : 'Net Billing',
    'Total Remaining Billing': 'Total Remaning Billing',
    'Current Month JTD'     : 'JTD Costs (B)',
    'Budgeted Gross Profit' : 'Budgeted Gross Proft ',
}

# ─────────────────────────────────────────────
# COLUMN FORMAT CLASSIFICATION
# ─────────────────────────────────────────────
# Accounting format:  $#,##0.00  with negatives as  ($#,##0.00)
ACCT_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

# Percentage format:  0.0%
PCT_FMT = '0.0%'

# Text columns get no special number format (General)
TEXT_COLS = {
    'Status', 'Job #', 'Job Name', 'Client Name',
    'Write Up / Write Down Potential ', 'Notes',
}

CURRENCY_COLS = {
    'As Sold Contract Amount',
    'Approved Change Orders',
    "CO's Added during WIP",
    'Current Contract Amount',
    'JTD Billings',
    'Retaininage',
    'Net Billing',
    '(Over) / Under Billings',
    'Total Remaning Billing',
    'Total Cost Projection (As of report date)',
    'Remaining Cost to Complete Projection (c)',
    'JTD Costs (B)',
    'Write Up / Write Down Amount',
    'Budgeted Gross Proft ',
    'Projected Gross Profit',
    'GP Variance (USD)',
    'JTD Revenue',
    'Unearned Rev',
    'Remaining Profit Margin',
    'JTD Cost',
    'JTD Profits',
    'YTD Revenue',
    'YTD Cost',
    'YTD Profit',
    'MTD Revenue',
    'MTD Cost',
    'MTD Profit',
    'YTD Cost.1',
    'YTD Revenue.1',
}

PERCENT_COLS = {
    'Retainage %',
    'Over / (Under) Billings %',
    'Billings %',
    '% Complete',
    'Marin %\nwriteup/wiritedown',
    'Projected Margin %',
    'GP Variance(%)',
}


def save_wip(df, path=None):
    from io import BytesIO
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, numbers

    if path is None:
        path = WIP_FILE
    out = df.copy()
    out.rename(columns=COL_RENAME_TO_WIP, inplace=True)
    for col in WIP_TARGET_COLS:
        if col not in out.columns:
            out[col] = 0.0
    extra_cols = [c for c in out.columns if c not in WIP_TARGET_COLS]
    out = out[WIP_TARGET_COLS + extra_cols]

    # --- Write raw data via pandas first ---
    out.to_excel(path, index=False)

    # --- Apply accounting formats via openpyxl ---
    wb = load_workbook(path)
    ws = wb.active

    # Build column-index → format map (1-indexed)
    col_names = [str(c.value) for c in ws[1]]
    for col_idx, col_name in enumerate(col_names, start=1):
        if col_name in CURRENCY_COLS:
            fmt = ACCT_FMT
        elif col_name in PERCENT_COLS:
            fmt = PCT_FMT
        else:
            continue
        # Apply to every data row (skip header row 1)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = fmt

    # Auto-fit column widths (rough heuristic)
    for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(col_name)) + 2
        if col_name in CURRENCY_COLS:
            max_len = max(max_len, 18)
        elif col_name in PERCENT_COLS:
            max_len = max(max_len, 10)
        ws.column_dimensions[col_letter].width = min(max_len, 28)

    # Bold header row
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    wb.save(path)

    # Also create download bytes with the same formatting
    buf = BytesIO()
    wb.save(buf)
    st.session_state['wip_download_bytes'] = buf.getvalue()
    return out

def wip_download_button(key="wip_dl"):
    if 'wip_download_bytes' in st.session_state:
        st.download_button(
            label="📥  Download WIP.xlsx",
            data=st.session_state['wip_download_bytes'],
            file_name="WIP.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=key,
        )


# ─────────────────────────────────────────────
# OVERHEAD JOBS & SHARED PIPELINE
# ─────────────────────────────────────────────
OVERHEAD_JOBS = {'25000', '26000'}

def run_jva_pipeline(file_obj):
    df_raw = pd.read_excel(file_obj)
    col_map = {
        'job_no'      : 'job_no',       'cost_code_no': 'cost_code_no',
        'costs_3'     : 'estimated_labor_cost',
        'costs_4'     : 'current_labor_cost_without_burden',
        'costs_5'     : 'estimated_material_cost',
        'costs_6'     : 'current_material_cost',
        'units_14'    : 'estimated_labor_hours',
        'units_15'    : 'current_labor_hours',
        'costs_8'     : 'current_labor_burden_cost',
        'costs_7'     : 'current_rented_equipment_cost',
        'costs_9'     : 'current_subcontractor_cost',
    }
    df = df_raw[list(col_map.keys())].rename(columns=col_map).copy()
    df['job_no']       = df['job_no'].astype(str).str.strip()
    df['cost_code_no'] = df['cost_code_no'].astype(str).str.strip()
    LABOR_SUFFIXES = ('FLD', 'FLM', 'PM')
    df['cost_class'] = df['cost_code_no'].apply(
        lambda x: 'Labor' if x.endswith(LABOR_SUFFIXES) else 'Material'
    )
    NUMERIC_COLS = [
        'estimated_labor_cost', 'current_labor_cost_without_burden',
        'estimated_material_cost', 'current_material_cost',
        'estimated_labor_hours', 'current_labor_hours',
        'current_labor_burden_cost', 'current_rented_equipment_cost',
        'current_subcontractor_cost',
    ]
    for col in NUMERIC_COLS:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    grouped = df.groupby(['job_no', 'cost_code_no', 'cost_class'], as_index=False)[NUMERIC_COLS].max()
    grouped['current_labor_cost']   = grouped['current_labor_cost_without_burden'] + grouped['current_labor_burden_cost']
    grouped['actual_material_cost'] = grouped['current_material_cost'] + grouped['current_rented_equipment_cost'] + grouped['current_subcontractor_cost']
    job_jtd = grouped.groupby('job_no', as_index=False).agg(
        cost_code_count   = ('cost_code_no',           'count'),
        est_labor_cost    = ('estimated_labor_cost',   'sum'),
        est_material_cost = ('estimated_material_cost','sum'),
        est_labor_hours   = ('estimated_labor_hours',  'sum'),
        act_labor_hours   = ('current_labor_hours',    'sum'),
        jtd_labor_cost    = ('current_labor_cost',     'sum'),
        jtd_material_cost = ('actual_material_cost',   'sum'),
    )
    job_jtd['jtd_total_cost']         = job_jtd['jtd_labor_cost'] + job_jtd['jtd_material_cost']
    job_jtd['labor_cost_variance']    = job_jtd['est_labor_cost']    - job_jtd['jtd_labor_cost']
    job_jtd['material_cost_variance'] = job_jtd['est_material_cost'] - job_jtd['jtd_material_cost']
    job_jtd['labor_hour_variance']    = job_jtd['est_labor_hours']   - job_jtd['act_labor_hours']
    return grouped, job_jtd, len(df_raw)

JVA_REQUIRED_COLS = [
    'job_no', 'cost_code_no', 'costs_3', 'costs_4', 'costs_5',
    'costs_6', 'costs_7', 'costs_8', 'costs_9', 'units_14', 'units_15',
]

def validate_jva(file_obj):
    df = pd.read_excel(file_obj, nrows=0)
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in JVA_REQUIRED_COLS if c not in df.columns]
    file_obj.seek(0)
    return missing

def validate_prev_wip(file_obj):
    df = pd.read_excel(file_obj)
    df.columns = [str(c).strip() for c in df.columns]
    remaining_col = 'Remaining Cost to Complete Projection (c)'
    if 'Job #' not in df.columns:
        return "Column **'Job #'** not found.", None
    if remaining_col not in df.columns:
        return f"Column **'{remaining_col}'** not found.", None
    df['Job #'] = df['Job #'].astype(str).str.strip()
    df[remaining_col] = pd.to_numeric(df[remaining_col], errors='coerce').fillna(0)
    return None, df

def process_jva_upload(file_obj, label):
    missing = validate_jva(file_obj)
    if missing:
        st.error(label + " is missing required columns: " + ', '.join(missing))
        return None
    try:
        grouped, job_jtd, raw_rows = run_jva_pipeline(file_obj)
        st.success(
            label + " loaded — "
            + "{:,}".format(raw_rows) + " raw rows → "
            + "{:,}".format(len(grouped)) + " cost-code rows → "
            + "{:,}".format(len(job_jtd)) + " jobs"
        )
        return grouped, job_jtd, raw_rows
    except Exception as e:
        st.error("Error processing " + label + ": " + str(e))
        return None


# ─────────────────────────────────────────────
# DATA LOADING ENGINE
# ─────────────────────────────────────────────
if 'master_wip' not in st.session_state:
    if os.path.exists(WIP_FILE):
        df_load = pd.read_excel(WIP_FILE)
    elif os.path.exists(MASTER_FILE):
        df_load = pd.read_excel(MASTER_FILE)
    else:
        df_load = pd.DataFrame(columns=ALL_COLS)
    df_load.columns = [str(c).strip() for c in df_load.columns]
    df_load.rename(columns={'Paula Biiling': 'Paula Billings'}, inplace=True)
    if 'Job #' in df_load.columns:
        df_load['Job #'] = df_load['Job #'].astype(str).str.strip()
    for col in ALL_COLS:
        if col not in df_load.columns:
            df_load[col] = 0.0 if any(x in col for x in ['Amount', 'Orders', '%', 'Billings', 'Retainage', 'Net']) else ""
    st.session_state.master_wip = df_load[ALL_COLS].copy()

if 'Paula Biiling' in st.session_state.master_wip.columns:
    st.session_state.master_wip.rename(columns={'Paula Biiling': 'Paula Billings'}, inplace=True)
for col in ALL_COLS:
    if col not in st.session_state.master_wip.columns:
        st.session_state.master_wip[col] = 0.0 if any(x in col for x in ['Amount', 'Orders', '%', 'Billings', 'Retainage', 'Net']) else ""

if os.path.exists(MASTER_FILE):
    df_master = pd.read_excel(MASTER_FILE)
    df_master.columns = [str(c).strip() for c in df_master.columns]
    df_master.rename(columns={'Paula Biiling': 'Paula Billings'}, inplace=True)
    if 'Paula Billings' in df_master.columns and 'Job #' in df_master.columns:
        df_master['Job #'] = df_master['Job #'].astype(str).str.strip()
        paula_map = df_master.set_index('Job #')['Paula Billings'].to_dict()
        st.session_state.master_wip['Paula Billings'] = st.session_state.master_wip['Job #'].map(paula_map).fillna(0)


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<div style="text-align:center;padding:1rem 0 0.5rem;">'
        '<span style="font-family:Montserrat,sans-serif;font-weight:800;font-size:1.4rem;">'
        '🏗️ Project Suite</span></div>',
        unsafe_allow_html=True,
    )
    total_projects  = len(st.session_state.master_wip)
    active_projects = len(st.session_state.master_wip[st.session_state.master_wip['Status'].astype(str).str.lower() == 'active'])
    st.markdown(
        f'<div style="text-align:center;font-size:0.8rem;opacity:0.7;margin-bottom:0.75rem;">'
        f'{total_projects} projects · {active_projects} active</div>',
        unsafe_allow_html=True,
    )
    st.markdown("---")
    page = st.radio(
        "NAVIGATION",
        ["📂 Project Registry", "📄 Contract Info", "💰 Billing Details",
         "📊 Cost Analysis", "📈 Projected Cost"],
        label_visibility="collapsed",
    )
    if page == "📄 Contract Info":
        st.markdown("---")
        st.markdown('<p style="font-size:0.75rem;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;opacity:0.6;margin-bottom:0.5rem;">Global Actions</p>', unsafe_allow_html=True)
        if st.button("🔄 Update All Totals", use_container_width=True):
            try:
                st.session_state.master_wip['Current Contract Amount'] = (
                    pd.to_numeric(st.session_state.master_wip['As Sold Contract Amount'], errors='coerce').fillna(0) +
                    pd.to_numeric(st.session_state.master_wip['Approved Change Orders'], errors='coerce').fillna(0) +
                    pd.to_numeric(st.session_state.master_wip["CO's Added during WIP"], errors='coerce').fillna(0)
                )
                save_wip(st.session_state.master_wip)
                st.success("All totals updated")
                st.rerun()
            except PermissionError:
                st.error("Close WIP.xlsx before updating.")
    if page == "💰 Billing Details":
        st.markdown("---")
        st.markdown('<p style="font-size:0.75rem;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;opacity:0.6;margin-bottom:0.5rem;">Global Actions</p>', unsafe_allow_html=True)
        if st.button("🔄 Recalculate All Billings", use_container_width=True):
            try:
                paula   = pd.to_numeric(st.session_state.master_wip['Paula Billings'], errors='coerce').fillna(0)
                jvm     = pd.to_numeric(st.session_state.master_wip['JVM Billings'],   errors='coerce').fillna(0)
                ret_pct = pd.to_numeric(st.session_state.master_wip['Retainage %'],    errors='coerce').fillna(0)
                st.session_state.master_wip['JTD Billings'] = paula + jvm
                st.session_state.master_wip['Retainage']   = ret_pct * (paula + jvm)
                st.session_state.master_wip['Net Billings'] = (paula + jvm) - st.session_state.master_wip['Retainage']
                save_wip(st.session_state.master_wip)
                st.success("All billing totals recalculated.")
                st.rerun()
            except PermissionError:
                st.error("Close WIP.xlsx before updating.")
    st.markdown("---")
    wip_download_button("dl_sidebar")
    st.markdown(f'<div style="text-align:center;font-size:0.7rem;opacity:0.45;padding-bottom:1rem;">💾 Saving to {WIP_FILE}</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────
# PAGE 1: PROJECT REGISTRY
# ─────────────────────────────────────────────
if page == "📂 Project Registry":
    page_header("📂", "Project Registry", f"{len(st.session_state.master_wip)} projects")
    df = st.session_state.master_wip
    metric_cards([
        ("Total Projects", str(len(df)), "navy"),
        ("Active",  str(len(df[df['Status'].astype(str).str.lower() == 'active'])),  "green"),
        ("Pending", str(len(df[df['Status'].astype(str).str.lower() == 'pending'])), ""),
        ("Closed",  str(len(df[df['Status'].astype(str).str.lower() == 'closed'])),  ""),
    ])
    job_list = ["-- Create New Project --"] + st.session_state.master_wip['Job #'].unique().tolist()
    selected_job_choice = st.selectbox("Select Job # to Edit or Create New", job_list)
    unique_statuses = st.session_state.master_wip['Status'].unique().tolist()
    status_options  = [str(s) for s in unique_statuses if pd.notnull(s) and str(s).strip() != ""]
    if not status_options:
        status_options = ["Active", "Pending", "Awarded", "Closed"]
    current_status = status_options[0]; current_job_name = ""; current_client_name = ""
    job_no_input = ""; is_edit_mode = False
    if selected_job_choice != "-- Create New Project --":
        idx = st.session_state.master_wip.index[st.session_state.master_wip['Job #'] == selected_job_choice][0]
        row = st.session_state.master_wip.iloc[idx]
        current_status      = str(row['Status'])
        current_job_name    = str(row['Job Name'])    if pd.notnull(row['Job Name'])    else ""
        current_client_name = str(row['Client Name']) if pd.notnull(row['Client Name']) else ""
        job_no_input = selected_job_choice; is_edit_mode = True
    with st.form("registry_form"):
        st.markdown(f'<p style="font-weight:600;color:#1e293b;">{"✏️ Editing" if is_edit_mode else "➕ New Project"}{"  ·  Job " + job_no_input if is_edit_mode else ""}</p>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            final_job_no = st.text_input("Job #",      value=job_no_input,        disabled=is_edit_mode)
            job_name     = st.text_input("Job Name",   value=current_job_name)
        with col2:
            status      = st.selectbox("Status", status_options, index=status_options.index(current_status) if current_status in status_options else 0)
            client_name = st.text_input("Client Name", value=current_client_name)
        if st.form_submit_button("💾  Save Project", use_container_width=True):
            mask = st.session_state.master_wip['Job #'] == final_job_no
            if mask.any():
                idx = st.session_state.master_wip.index[mask][0]
                st.session_state.master_wip.at[idx, 'Status']      = status
                st.session_state.master_wip.at[idx, 'Job Name']    = job_name
                st.session_state.master_wip.at[idx, 'Client Name'] = client_name
            else:
                new_row = {c: (0.0 if any(x in c for x in ['Amount','Orders','%','Billings','Retainage','Net']) else "") for c in ALL_COLS}
                new_row.update({'Status': status, 'Job #': final_job_no, 'Job Name': job_name, 'Client Name': client_name})
                st.session_state.master_wip = pd.concat([st.session_state.master_wip, pd.DataFrame([new_row])], ignore_index=True)
            save_wip(st.session_state.master_wip)
            st.rerun()
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown("##### All Projects")
    st.dataframe(st.session_state.master_wip[P1_COLS], use_container_width=True, hide_index=True, height=min(400, 56+35*len(st.session_state.master_wip)))


# ─────────────────────────────────────────────
# PAGE 2: CONTRACT INFO
# ─────────────────────────────────────────────
elif page == "📄 Contract Info":
    page_header("📄", "Contract Information")
    if st.session_state.master_wip.empty:
        st.warning("No projects found.")
    else:
        job_list     = st.session_state.master_wip['Job #'].tolist()
        selected_job = st.selectbox("Select Job #", job_list)
        idx          = st.session_state.master_wip.index[st.session_state.master_wip['Job #'] == selected_job][0]
        row          = st.session_state.master_wip.iloc[idx]
        as_sold      = float(pd.to_numeric(row['As Sold Contract Amount'], errors='coerce') or 0)
        approved_co  = float(pd.to_numeric(row['Approved Change Orders'],  errors='coerce') or 0)
        wip_co       = float(pd.to_numeric(row["CO's Added during WIP"],   errors='coerce') or 0)
        current_total= float(pd.to_numeric(row['Current Contract Amount'], errors='coerce') or 0)
        metric_cards([
            ("As Sold",          f"${as_sold:,.2f}",       "navy"),
            ("Approved CO's",    f"${approved_co:,.2f}",   ""),
            ("WIP CO's",         f"${wip_co:,.2f}",        ""),
            ("Current Contract", f"${current_total:,.2f}", "green"),
        ])
        with st.form("contract_form"):
            st.markdown(f'<p style="font-weight:600;color:#1e293b;">Financials · Job {selected_job}</p>', unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            with c1:
                st.text_input("As Sold Contract Amount", value=f"{row['As Sold Contract Amount']:.2f}", disabled=True)
                raw_dt     = row['Contract Date']
                default_dt = pd.to_datetime(raw_dt).date() if pd.notnull(raw_dt) else date.today()
                st.date_input("Contract Date", value=default_dt, disabled=True)
            with c2:
                app_co_str = st.text_input("Approved Change Orders", value=f"{row['Approved Change Orders']:.2f}")
                ret_raw    = pd.to_numeric(row['Retainage %'], errors='coerce') or 0
                st.text_input("Retainage %", value=f"{ret_raw*100:.1f}%", disabled=True)
            with c3:
                wip_co_key = "CO's Added during WIP"
                wip_co_str = st.text_input("CO's Added during WIP", value=f"{row[wip_co_key]:.2f}")
                notes      = st.text_area("Notes", value=str(row['Notes']) if pd.notnull(row['Notes']) else "")
            b1, b2 = st.columns(2)
            update_job = b1.form_submit_button("💾  Save Changes")
            calc_job   = b2.form_submit_button("🧮  Calculate Total")
            if update_job or calc_job:
                try:
                    as_sold_v = float(row['As Sold Contract Amount'])
                    app_co_v  = float(app_co_str)
                    wip_co_v  = float(wip_co_str)
                    total     = as_sold_v + app_co_v + wip_co_v if calc_job else row['Current Contract Amount']
                    st.session_state.master_wip.loc[idx, ['Approved Change Orders', "CO's Added during WIP", 'Current Contract Amount', 'Notes']] = [app_co_v, wip_co_v, total, notes]
                    save_wip(st.session_state.master_wip)
                    st.rerun()
                except ValueError:
                    st.error("Check your number formats.")
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown("##### Contract Overview")
    st.dataframe(st.session_state.master_wip[['Job #','Job Name']+P2_COLS], use_container_width=True, hide_index=True, height=min(400, 56+35*len(st.session_state.master_wip)))


# ─────────────────────────────────────────────
# PAGE 3: BILLING DETAILS
# ─────────────────────────────────────────────
elif page == "💰 Billing Details":
    page_header("💰", "Billing Details")
    if st.session_state.master_wip.empty:
        st.warning("No projects found.")
    else:
        st.markdown('<p style="font-weight:700;color:#1e293b;font-size:1.05rem;">📤 Import JVM Billings</p><p style="font-size:0.84rem;color:#64748b;">Upload a CSV with columns <b>job_no</b> and <b>JVM Billings</b>.</p>', unsafe_allow_html=True)
        uploaded_csv = st.file_uploader("Upload JVM Billings CSV", type=["csv"], label_visibility="collapsed")
        if uploaded_csv is not None:
            try:
                jvm_df = pd.read_csv(uploaded_csv)
                jvm_df.columns = [str(c).strip() for c in jvm_df.columns]
                if 'job_no' not in jvm_df.columns or 'JVM Billings' not in jvm_df.columns:
                    st.error("CSV must contain columns **job_no** and **JVM Billings**.")
                else:
                    jvm_df['job_no'] = jvm_df['job_no'].astype(str).str.strip()
                    jvm_cleaned = jvm_df['JVM Billings'].astype(str).str.strip().str.replace('$','',regex=False).str.replace(',','',regex=False).str.strip().replace({'-':'0','':'0'})
                    jvm_df['JVM Billings'] = pd.to_numeric(jvm_cleaned, errors='coerce').fillna(0)
                    with st.expander("Preview uploaded data", expanded=True):
                        st.dataframe(jvm_df, use_container_width=True, hide_index=True)
                    if st.button("✅  Apply JVM Billings to WIP", use_container_width=True):
                        matched = 0; unmatched = []
                        for _, csv_row in jvm_df.iterrows():
                            mask = st.session_state.master_wip['Job #'] == csv_row['job_no']
                            if mask.any():
                                st.session_state.master_wip.at[st.session_state.master_wip.index[mask][0], 'JVM Billings'] = csv_row['JVM Billings']
                                matched += 1
                            else:
                                unmatched.append(csv_row['job_no'])
                        paula   = pd.to_numeric(st.session_state.master_wip['Paula Billings'], errors='coerce').fillna(0)
                        jvm     = pd.to_numeric(st.session_state.master_wip['JVM Billings'],   errors='coerce').fillna(0)
                        ret_pct = pd.to_numeric(st.session_state.master_wip['Retainage %'],    errors='coerce').fillna(0)
                        st.session_state.master_wip['JTD Billings'] = paula + jvm
                        st.session_state.master_wip['Retainage']    = ret_pct * (paula + jvm)
                        st.session_state.master_wip['Net Billings'] = (paula + jvm) - st.session_state.master_wip['Retainage']
                        save_wip(st.session_state.master_wip)
                        st.success(f"Done — **{matched}** job(s) updated.")
                        if unmatched:
                            st.warning("Unmatched Job #s: " + ', '.join(unmatched))
                        st.rerun()
            except Exception as e:
                st.error(f"Error reading CSV: {e}")
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
        st.markdown('<p style="font-weight:700;color:#1e293b;font-size:1.05rem;">🔍 Job Billing Details</p>', unsafe_allow_html=True)
        job_list     = st.session_state.master_wip['Job #'].tolist()
        selected_job = st.selectbox("Select Job #", job_list, key="billing_job_select")
        idx          = st.session_state.master_wip.index[st.session_state.master_wip['Job #'] == selected_job][0]
        row          = st.session_state.master_wip.iloc[idx]
        paula_val    = float(pd.to_numeric(row['Paula Billings'], errors='coerce') or 0)
        jvm_val      = float(pd.to_numeric(row['JVM Billings'],   errors='coerce') or 0)
        jtd_val      = paula_val + jvm_val
        ret_pct_val  = float(pd.to_numeric(row['Retainage %'],    errors='coerce') or 0)
        retainage_val= ret_pct_val * jtd_val
        net_val      = jtd_val - retainage_val
        metric_cards([("Paula Billings", f"${paula_val:,.2f}", "navy"), ("JVM Billings", f"${jvm_val:,.2f}", "navy"), ("JTD Billings", f"${jtd_val:,.2f}", "")])
        metric_cards([("Retainage %", f"{ret_pct_val*100:.1f}%", "red"), ("Retainage $", f"${retainage_val:,.2f}", "red"), ("Net Billings", f"${net_val:,.2f}", "green")])
        if st.button("🧮  Recalculate This Job", use_container_width=True, key="recalc_billing"):
            st.session_state.master_wip.at[idx, 'JTD Billings'] = jtd_val
            st.session_state.master_wip.at[idx, 'Retainage']    = retainage_val
            st.session_state.master_wip.at[idx, 'Net Billings']  = net_val
            save_wip(st.session_state.master_wip)
            st.success(f"Job {selected_job} recalculated and saved.")
            st.rerun()
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
        st.markdown("##### Billing Overview — All Jobs")
        display_df = st.session_state.master_wip[['Job #','Job Name','Paula Billings','JVM Billings','JTD Billings','Retainage %','Retainage','Net Billings']].copy()
        for col in ['Paula Billings','JVM Billings','JTD Billings','Retainage','Net Billings']:
            display_df[col] = pd.to_numeric(display_df[col], errors='coerce').fillna(0)
        display_df['Retainage %'] = pd.to_numeric(display_df['Retainage %'], errors='coerce').fillna(0) * 100
        st.dataframe(display_df.style.format({'Paula Billings':'${:,.2f}','JVM Billings':'${:,.2f}','JTD Billings':'${:,.2f}','Retainage':'${:,.2f}','Net Billings':'${:,.2f}','Retainage %':'{:.1f}%'}), use_container_width=True, hide_index=True, height=min(450, 56+35*len(st.session_state.master_wip)))


# ─────────────────────────────────────────────
# PAGE 4: COST ANALYSIS
# ─────────────────────────────────────────────
elif page == "📊 Cost Analysis":
    page_header("📊", "Cost Analysis")

    st.markdown(
        '<p style="font-weight:700;color:#1e293b;font-size:1.05rem;margin-bottom:0.25rem;">📤 Upload JVA Cost Files</p>'
        '<p style="font-size:0.84rem;color:#64748b;margin-bottom:0.75rem;">'
        'Upload all JVA files here. <b>Current Period</b> is this month\'s cumulative snapshot. '
        '<b>Prior Period</b> (start of year) is used for YTD cost and overhead reallocation. '
        '<b>Previous Month-End</b> is used on the Projected Cost page for monthly delta.</p>',
        unsafe_allow_html=True,
    )

    col_up1, col_up2, col_up3 = st.columns(3)
    with col_up1:
        st.markdown('<p style="font-size:0.85rem;font-weight:600;color:#1e293b;">📄 Current Period JVA</p>', unsafe_allow_html=True)
        uploaded_current    = st.file_uploader("Current JVA",       type=["xlsx","xls"], label_visibility="collapsed", key="jva_current")
    with col_up2:
        st.markdown('<p style="font-size:0.85rem;font-weight:600;color:#1e293b;">📄 Prior Period JVA (for YTD)</p>', unsafe_allow_html=True)
        uploaded_prior      = st.file_uploader("Prior JVA",         type=["xlsx","xls"], label_visibility="collapsed", key="jva_prior")
    with col_up3:
        st.markdown('<p style="font-size:0.85rem;font-weight:600;color:#1e293b;">📄 Previous Month-End JVA</p>', unsafe_allow_html=True)
        uploaded_prev_month = st.file_uploader("Previous Month JVA",type=["xlsx","xls"], label_visibility="collapsed", key="jva_prev_month")

    if uploaded_current is not None:
        result = process_jva_upload(uploaded_current, "Current Period JVA")
        if result:
            grouped, job_jtd, _ = result
            st.session_state['jva_grouped'] = grouped
            st.session_state['jva_job_jtd'] = job_jtd

    if uploaded_prior is not None:
        result = process_jva_upload(uploaded_prior, "Prior Period JVA")
        if result:
            _, job_jtd_prior, _ = result
            st.session_state['jva_job_jtd_prior'] = job_jtd_prior

    if uploaded_prev_month is not None:
        result = process_jva_upload(uploaded_prev_month, "Previous Month-End JVA")
        if result:
            _, prev_month_jtd, _ = result
            st.session_state['jva_prev_month_jtd'] = prev_month_jtd

    if 'jva_job_jtd' not in st.session_state:
        st.info("Upload the Current Period JVA file above to begin Cost Analysis.")
        st.stop()

    grouped = st.session_state['jva_grouped']
    job_jtd = st.session_state['jva_job_jtd'].copy()

    prior_available = 'jva_job_jtd_prior' in st.session_state
    if prior_available:
        prior_map = st.session_state['jva_job_jtd_prior'].set_index('job_no')['jtd_total_cost'].to_dict()
        job_jtd['ytd_cost'] = job_jtd.apply(lambda r: r['jtd_total_cost'] - prior_map.get(r['job_no'], 0), axis=1)
    else:
        job_jtd['ytd_cost'] = None

    overhead_mask = job_jtd['job_no'].isin(OVERHEAD_JOBS)
    other_mask    = ~overhead_mask
    realloc_pool  = job_jtd.loc[overhead_mask, 'jtd_total_cost'].sum()
    job_jtd['adjusted_jtd_cost'] = job_jtd['jtd_total_cost'].copy()

    if prior_available and realloc_pool != 0:
        other_df      = job_jtd.loc[other_mask].copy()
        positive_ytd  = other_df['ytd_cost'].fillna(0)
        eligible_mask = positive_ytd > 0
        ratio_base    = positive_ytd[eligible_mask].sum()
        if ratio_base > 0:
            ratios       = positive_ytd[eligible_mask] / ratio_base
            allocation   = ratios * realloc_pool
            eligible_idx = other_df.index[eligible_mask]
            job_jtd.loc[eligible_idx, 'adjusted_jtd_cost'] = job_jtd.loc[eligible_idx, 'jtd_total_cost'] + allocation

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown('<p style="font-weight:700;color:#1e293b;font-size:1.05rem;margin-bottom:0.5rem;">🔍 Job Cost Detail</p>', unsafe_allow_html=True)
    job_options  = sorted(job_jtd['job_no'].unique())
    selected_job = st.selectbox("Select Job #", job_options, key="cost_job_select")
    job_row      = job_jtd[job_jtd['job_no'] == selected_job].iloc[0]

    def render_all_cards(job_row):
        COLOR = {"grey":"#94a3b8","green":"#16a34a","amber":"#d97706","red":"#D10D38","navy":"#153275","slate":"#1e293b"}
        LABEL = "font-size:0.72rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.15rem;"
        CARD  = "flex:1;background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:1.1rem 1rem;text-align:center;box-sizing:border-box;display:flex;flex-direction:column;justify-content:center;align-items:center;gap:0.18rem;"
        ROW   = "display:flex;gap:0.85rem;margin-bottom:0.85rem;align-items:stretch;"
        def pick_color(actual, estimated):
            if estimated == 0: return COLOR["grey"]
            r = actual / estimated
            if r <= 0.80: return COLOR["green"]
            elif r <= 1.00: return COLOR["amber"]
            else: return COLOR["red"]
        def vcrd(label, est, act, var, unit="$"):
            col = pick_color(act, est); sign = "+" if var >= 0 else "-"; vc = COLOR["green"] if var >= 0 else COLOR["red"]
            es = "${:,.2f}".format(est) if unit=="$" else "{:,.1f} hrs".format(est)
            ac = "${:,.2f}".format(act) if unit=="$" else "{:,.1f} hrs".format(act)
            va = "${:,.2f}".format(abs(var)) if unit=="$" else "{:,.1f} hrs".format(abs(var))
            return '<div style="'+CARD+'"><div style="'+LABEL+'">'+label+'</div><div style="font-size:0.75rem;color:#64748b;">Est: '+es+'</div><div style="font-family:Montserrat,sans-serif;font-size:1.15rem;font-weight:700;color:'+col+';">'+ac+'</div><div style="font-size:0.75rem;font-weight:600;color:'+vc+';">'+sign+va+' variance</div></div>'
        def pcrd(label, value, col):
            return '<div style="'+CARD+'"><div style="'+LABEL+'">'+label+'</div><div style="font-family:Montserrat,sans-serif;font-size:1.25rem;font-weight:700;color:'+col+';">'+value+'</div></div>'
        r1c1 = vcrd("Labor Cost",    float(job_row['est_labor_cost']),    float(job_row['jtd_labor_cost']),    float(job_row['labor_cost_variance']),    unit="$")
        r1c2 = vcrd("Material Cost", float(job_row['est_material_cost']), float(job_row['jtd_material_cost']), float(job_row['material_cost_variance']), unit="$")
        r1c3 = vcrd("Labor Hours",   float(job_row['est_labor_hours']),   float(job_row['act_labor_hours']),   float(job_row['labor_hour_variance']),   unit="hrs")
        r1c4 = pcrd("JTD Total Cost","${:,.2f}".format(float(job_row['jtd_total_cost'])), COLOR["green"])
        r2c1 = pcrd("JTD Labor Cost","${:,.2f}".format(float(job_row['jtd_labor_cost'])), COLOR["navy"])
        r2c2 = pcrd("JTD Material Cost","${:,.2f}".format(float(job_row['jtd_material_cost'])), COLOR["navy"])
        ytd_raw = job_row.get('ytd_cost', None)
        r2c3 = pcrd("YTD Cost", "${:,.2f}".format(float(ytd_raw)), COLOR["slate"]) if ytd_raw is not None and not pd.isna(ytd_raw) else pcrd("YTD Cost", "Upload prior period", COLOR["grey"])
        adj_jtd = job_row.get('adjusted_jtd_cost', None)
        r2c4 = pcrd("Adjusted JTD Cost", "${:,.2f}".format(float(adj_jtd)), COLOR["amber"]) if adj_jtd is not None and not pd.isna(adj_jtd) else pcrd("Adjusted JTD Cost", "Upload prior period", COLOR["grey"])
        st.markdown('<div style="'+ROW+'">'+r1c1+r1c2+r1c3+r1c4+'</div>'+'<div style="'+ROW+'">'+r2c1+r2c2+r2c3+r2c4+'</div>', unsafe_allow_html=True)

    render_all_cards(job_row)
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    st.markdown('<p style="font-weight:700;color:#1e293b;font-size:1.05rem;margin-bottom:0.5rem;">📑 WIP Register — Calculated Columns</p>', unsafe_allow_html=True)
    wip_display = st.session_state.master_wip.copy()
    wip_display['_job_key'] = wip_display['Job #'].astype(str).str.strip()
    wip_display = wip_display[~wip_display['_job_key'].isin(OVERHEAD_JOBS)].drop(columns=['_job_key']).reset_index(drop=True)

    jtd_map = job_jtd.set_index('job_no')['jtd_total_cost'].to_dict()
    wip_display['JTD Total Cost'] = wip_display['Job #'].map(lambda j: jtd_map.get(str(j).strip(), 0)).fillna(0)

    ytd_map = job_jtd.set_index('job_no')['ytd_cost'].to_dict()
    wip_display['YTD Cost'] = wip_display['Job #'].map(lambda j: ytd_map.get(str(j).strip(), None))

    adj_map = job_jtd.set_index('job_no')['adjusted_jtd_cost'].to_dict()
    wip_display['Adjusted JTD Cost'] = wip_display['Job #'].map(lambda j: adj_map.get(str(j).strip(), None))

    wip_display['Total Budgeted Cost (a)'] = 0.0
    if os.path.exists(MASTER_FILE):
        df_mb = pd.read_excel(MASTER_FILE)
        df_mb.columns = [str(c).strip() for c in df_mb.columns]
        df_mb['Job #'] = df_mb['Job #'].astype(str).str.strip()
        if 'Total Budgeted Cost (a)' in df_mb.columns:
            budget_map = df_mb.set_index('Job #')['Total Budgeted Cost (a)'].to_dict()
            wip_display['Total Budgeted Cost (a)'] = wip_display['Job #'].map(lambda j: budget_map.get(str(j).strip(), 0)).fillna(0)
        else:
            st.warning("Column 'Total Budgeted Cost (a)' not found in Master_Project_Registry.xlsx.")
    else:
        st.warning("Master_Project_Registry.xlsx not found.")

    wip_show_cols = [c for c in ['Job #','Job Name','Status','Current Contract Amount','Total Budgeted Cost (a)','JTD Total Cost','YTD Cost','Adjusted JTD Cost','% Complete','JTD Billings','(Over) / Under Billings','Over / (Under) Billings %','Net Billings'] if c in wip_display.columns]
    fmt_dict = {c:'${:,.2f}' for c in ['Current Contract Amount','Total Budgeted Cost (a)','JTD Total Cost','YTD Cost','Adjusted JTD Cost','JTD Billings','(Over) / Under Billings','Net Billings'] if c in wip_show_cols}
    if '% Complete' in wip_show_cols: fmt_dict['% Complete'] = '{:.1%}'
    if 'Over / (Under) Billings %' in wip_show_cols: fmt_dict['Over / (Under) Billings %'] = '{:.1%}'
    st.dataframe(wip_display[wip_show_cols].style.format(fmt_dict), use_container_width=True, hide_index=True, height=min(500, 56+35*len(wip_display)))

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("💾  Calculate & Save WIP to WIP.xlsx", use_container_width=True):
        try:
            adj_jtd  = pd.to_numeric(wip_display['Adjusted JTD Cost'],      errors='coerce').fillna(0)
            budget   = pd.to_numeric(wip_display['Total Budgeted Cost (a)'], errors='coerce').fillna(0)
            contract = pd.to_numeric(wip_display['Current Contract Amount'], errors='coerce').fillna(0)
            wip_display['% Complete']   = (adj_jtd / contract.where(contract != 0, other=1)).where(contract != 0, other=0)
            wip_display['JTD Billings'] = pd.to_numeric(wip_display['Paula Billings'], errors='coerce').fillna(0) + pd.to_numeric(wip_display['JVM Billings'], errors='coerce').fillna(0)
            pct_comp  = pd.to_numeric(wip_display['% Complete'],   errors='coerce').fillna(0)
            jtd_bill  = pd.to_numeric(wip_display['JTD Billings'], errors='coerce').fillna(0)
            wip_display['(Over) / Under Billings']   = (contract * pct_comp) - jtd_bill
            over_under = pd.to_numeric(wip_display['(Over) / Under Billings'], errors='coerce').fillna(0)
            wip_display['Over / (Under) Billings %'] = (over_under / contract.where(contract != 0, other=1)).where(contract != 0, other=0)
            save_wip(wip_display)
            st.session_state.master_wip = wip_display.copy()
            st.success("WIP calculated and saved (overhead jobs 25000/26000 excluded).")
            wip_download_button("dl_cost_analysis")
        except PermissionError:
            st.error("Close WIP.xlsx before saving.")
        except Exception as e:
            st.error("Save failed: " + str(e))


# ─────────────────────────────────────────────
# PAGE 5: PROJECTED COST  (streamlined layout)
# ─────────────────────────────────────────────
elif page == "📈 Projected Cost":
    page_header("📈", "Projected Cost")

    st.markdown(
        '<p style="font-size:0.84rem;color:#64748b;margin-bottom:0.75rem;">'
        'Calculates <b>Remaining Cost to Complete Projection (c)</b> and all GP / revenue metrics.</p>',
        unsafe_allow_html=True,
    )

    if 'jva_job_jtd' not in st.session_state:
        st.warning("⚠️ Go to **📊 Cost Analysis** and upload the Current Period JVA first.")
        st.stop()
    if 'jva_prev_month_jtd' not in st.session_state:
        st.warning("⚠️ Go to **📊 Cost Analysis** and upload the Previous Month-End JVA first.")
        st.stop()

    st.markdown('<p style="font-weight:700;color:#1e293b;font-size:1.05rem;margin-bottom:0.5rem;">📤 Upload Previous Month WIP</p>', unsafe_allow_html=True)
    uploaded_prev_wip = st.file_uploader("Previous Month WIP", type=["xlsx","xls"], label_visibility="collapsed", key="proj_prev_wip")
    if uploaded_prev_wip is not None:
        err, prev_wip_df = validate_prev_wip(uploaded_prev_wip)
        if err:
            st.error(err)
        else:
            st.session_state['proj_prev_wip_data'] = prev_wip_df
            st.success("Previous month WIP loaded — " + "{:,}".format(len(prev_wip_df)) + " rows.")

    if 'proj_prev_wip_data' not in st.session_state:
        st.info("Upload the Previous Month WIP above to calculate projections.")
        st.stop()

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    remaining_col            = 'Remaining Cost to Complete Projection (c)'
    total_cost_proj_col      = 'Total Cost Projection (As of report date)'
    total_remaining_bill_col = 'Total Remaining Billing'
    billings_pct_col         = 'Billings %'
    jtd_revenue_col          = 'JTD Revenue'

    current_jtd_map  = st.session_state['jva_job_jtd'].set_index('job_no')['jtd_total_cost'].to_dict()
    prev_jtd_map     = st.session_state['jva_prev_month_jtd'].set_index('job_no')['jtd_total_cost'].to_dict()
    prev_remaining_map = st.session_state['proj_prev_wip_data'].set_index('Job #')[remaining_col].to_dict()

    proj_display = st.session_state.master_wip.copy()
    proj_display['_job_key'] = proj_display['Job #'].astype(str).str.strip()
    proj_display = proj_display[~proj_display['_job_key'].isin(OVERHEAD_JOBS)].drop(columns=['_job_key']).reset_index(drop=True)

    proj_display['Current Month JTD']    = proj_display['Job #'].map(lambda j: current_jtd_map.get(str(j).strip(), 0)).fillna(0)
    proj_display['Previous Month JTD']   = proj_display['Job #'].map(lambda j: prev_jtd_map.get(str(j).strip(), 0)).fillna(0)
    proj_display['Previous Remaining (c)']= proj_display['Job #'].map(lambda j: prev_remaining_map.get(str(j).strip(), 0)).fillna(0)

    proj_display['Delta (Current - Prior JTD)'] = proj_display['Current Month JTD'] - proj_display['Previous Month JTD']
    proj_display[remaining_col]                  = proj_display['Previous Remaining (c)'] - proj_display['Delta (Current - Prior JTD)']

    jtd_cost = pd.to_numeric(proj_display['Current Month JTD'],       errors='coerce').fillna(0)
    remain_c = pd.to_numeric(proj_display[remaining_col],             errors='coerce').fillna(0)
    contract = pd.to_numeric(proj_display['Current Contract Amount'], errors='coerce').fillna(0)

    proj_display['JTD Billings'] = (
        pd.to_numeric(proj_display['Paula Billings'], errors='coerce').fillna(0) +
        pd.to_numeric(proj_display['JVM Billings'],   errors='coerce').fillna(0)
    )
    jtd_bill = pd.to_numeric(proj_display['JTD Billings'], errors='coerce').fillna(0)

    adj_jtd_map = st.session_state['jva_job_jtd'].set_index('job_no')['jtd_total_cost'].to_dict()
    proj_display['Adjusted JTD Cost'] = proj_display['Job #'].map(lambda j: adj_jtd_map.get(str(j).strip(), 0)).fillna(0)

    proj_display['Total Budgeted Cost (a)'] = 0.0
    if os.path.exists(MASTER_FILE):
        df_mb = pd.read_excel(MASTER_FILE)
        df_mb.columns = [str(c).strip() for c in df_mb.columns]
        df_mb['Job #'] = df_mb['Job #'].astype(str).str.strip()
        if 'Total Budgeted Cost (a)' in df_mb.columns:
            budget_map = df_mb.set_index('Job #')['Total Budgeted Cost (a)'].to_dict()
            proj_display['Total Budgeted Cost (a)'] = proj_display['Job #'].map(lambda j: budget_map.get(str(j).strip(), 0)).fillna(0)

    adj_jtd      = pd.to_numeric(proj_display['Adjusted JTD Cost'],      errors='coerce').fillna(0)
    budget       = pd.to_numeric(proj_display['Total Budgeted Cost (a)'], errors='coerce').fillna(0)
    pct_complete = (adj_jtd / contract.where(contract != 0, other=1)).where(contract != 0, other=0)
    proj_display['% Complete'] = pct_complete

    proj_display[total_cost_proj_col]      = jtd_cost + remain_c
    proj_display[total_remaining_bill_col] = contract - jtd_bill
    proj_display[billings_pct_col]         = (jtd_bill / contract.where(contract != 0, other=1)).where(contract != 0, other=0)
    proj_display[jtd_revenue_col]          = contract * pct_complete
    proj_display['(Over) / Under Billings'] = (contract * pct_complete) - jtd_bill
    over_under = pd.to_numeric(proj_display['(Over) / Under Billings'], errors='coerce').fillna(0)
    proj_display['Over / (Under) Billings %'] = (over_under / contract.where(contract != 0, other=1)).where(contract != 0, other=0)
    ret_pct = pd.to_numeric(proj_display['Retainage %'], errors='coerce').fillna(0)
    proj_display['Retainage']    = ret_pct * jtd_bill
    proj_display['Net Billings'] = jtd_bill - proj_display['Retainage']

    budgeted_margin = (1 - budget / contract.where(contract != 0, other=1)).where(contract != 0, other=0)
    proj_display['Budgeted Job Margin %']  = budgeted_margin
    proj_display['Budgeted Gross Profit']  = budgeted_margin * contract

    total_cost_proj_vals = pd.to_numeric(proj_display[total_cost_proj_col], errors='coerce').fillna(0)
    projected_margin = (1 - total_cost_proj_vals / contract.where(contract != 0, other=1)).where(contract != 0, other=0)
    proj_display['Projected Gross Margin %'] = projected_margin
    proj_display['Projected Gross Profit']   = projected_margin * contract

    proj_display['GP Variance(%)']   = budgeted_margin - projected_margin
    proj_display['GP Variance (USD)'] = (
        pd.to_numeric(proj_display['Projected Gross Profit'], errors='coerce').fillna(0) -
        pd.to_numeric(proj_display['Budgeted Gross Profit'],  errors='coerce').fillna(0)
    )
    proj_display['Unearned Rev'] = contract - pd.to_numeric(proj_display[jtd_revenue_col], errors='coerce').fillna(0)

    # ── Streamlined per-job KPI view ──
    st.markdown('<p style="font-weight:700;color:#1e293b;font-size:1.05rem;margin-bottom:0.5rem;">🔍 Job Projection Detail</p>', unsafe_allow_html=True)
    job_options       = sorted(proj_display['Job #'].unique())
    selected_proj_job = st.selectbox("Select Job # to inspect", job_options, key="proj_job_select")
    prow              = proj_display[proj_display['Job #'] == selected_proj_job].iloc[0]

    # ── Compact KPI table instead of 6 rows of metric cards ──
    def fmt_acct(v):
        """Format a value as accounting: $1,234.56 or ($1,234.56)"""
        val = float(v) if pd.notnull(v) else 0.0
        if val < 0:
            return f'<span style="color:#D10D38;">(${abs(val):,.2f})</span>'
        return f'${val:,.2f}'

    def fmt_pct(v):
        """Format a value as percentage: 12.3%"""
        val = float(v) if pd.notnull(v) else 0.0
        if val < 0:
            return f'<span style="color:#D10D38;">({abs(val)*100:.1f}%)</span>'
        return f'{val*100:.1f}%'

    # Build structured KPI sections
    kpi_sections = [
        {
            "title": "Cost Projection",
            "rows": [
                ("Current Contract Amount",    fmt_acct(prow['Current Contract Amount'])),
                ("JTD Cost (Current Period)",   fmt_acct(prow['Current Month JTD'])),
                ("Remaining Cost (c)",          fmt_acct(prow[remaining_col])),
                ("Total Cost Projection",       fmt_acct(prow[total_cost_proj_col])),
                ("% Complete",                  fmt_pct(prow['% Complete'])),
            ],
        },
        {
            "title": "Billing & Revenue",
            "rows": [
                ("JTD Billings",                fmt_acct(prow['JTD Billings'])),
                ("Total Remaining Billing",     fmt_acct(prow[total_remaining_bill_col])),
                ("Billings %",                  fmt_pct(prow[billings_pct_col])),
                ("JTD Revenue",                 fmt_acct(prow[jtd_revenue_col])),
                ("Unearned Revenue",            fmt_acct(prow['Unearned Rev'])),
                ("(Over) / Under Billings",     fmt_acct(prow['(Over) / Under Billings'])),
            ],
        },
        {
            "title": "Gross Profit Analysis",
            "rows": [
                ("Budgeted Gross Profit",       fmt_acct(prow['Budgeted Gross Profit'])),
                ("Projected Gross Profit",      fmt_acct(prow['Projected Gross Profit'])),
                ("GP Variance (USD)",           fmt_acct(prow['GP Variance (USD)'])),
                ("Budgeted Job Margin %",       fmt_pct(prow['Budgeted Job Margin %'])),
                ("Projected Gross Margin %",    fmt_pct(prow['Projected Gross Margin %'])),
                ("GP Variance %",               fmt_pct(prow['GP Variance(%)'])),
            ],
        },
    ]

    # Render as three side-by-side tables
    cols = st.columns(3)
    for i, section in enumerate(kpi_sections):
        with cols[i]:
            tbl_rows = ""
            for label, value in section["rows"]:
                tbl_rows += (
                    f'<tr>'
                    f'<td style="padding:0.45rem 0.6rem;font-size:0.82rem;color:#475569;border-bottom:1px solid #f1f5f9;">{label}</td>'
                    f'<td style="padding:0.45rem 0.6rem;font-size:0.88rem;font-weight:600;color:#1e293b;text-align:right;border-bottom:1px solid #f1f5f9;font-family:Montserrat,sans-serif;">{value}</td>'
                    f'</tr>'
                )
            st.markdown(
                f'<div style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;overflow:hidden;">'
                f'<div style="background:linear-gradient(135deg,#153275,#1a3f8f);color:#fff;padding:0.6rem 0.8rem;font-family:Montserrat,sans-serif;font-weight:700;font-size:0.82rem;text-transform:uppercase;letter-spacing:0.05em;">{section["title"]}</div>'
                f'<table style="width:100%;border-collapse:collapse;">{tbl_rows}</table>'
                f'</div>',
                unsafe_allow_html=True,
            )

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown("##### Projection — All Jobs")

    proj_show_cols = [c for c in [
        'Job #', 'Job Name', 'Status',
        'Current Month JTD', 'Previous Month JTD', 'Delta (Current - Prior JTD)',
        'Previous Remaining (c)', remaining_col, total_cost_proj_col,
        'Current Contract Amount', 'Total Budgeted Cost (a)',
        'JTD Billings', total_remaining_bill_col, billings_pct_col,
        '% Complete', jtd_revenue_col, 'Unearned Rev',
        'Budgeted Job Margin %', 'Budgeted Gross Profit',
        'Projected Gross Margin %', 'Projected Gross Profit',
        'GP Variance(%)', 'GP Variance (USD)',
    ] if c in proj_display.columns]

    proj_fmt = {}
    for c in ['Current Month JTD','Previous Month JTD','Delta (Current - Prior JTD)',
              'Previous Remaining (c)', remaining_col, total_cost_proj_col,
              'Current Contract Amount','Total Budgeted Cost (a)','JTD Billings',
              total_remaining_bill_col, jtd_revenue_col,
              'Unearned Rev','Budgeted Gross Profit','Projected Gross Profit','GP Variance (USD)']:
        if c in proj_show_cols: proj_fmt[c] = '${:,.2f}'
    for c in [billings_pct_col, '% Complete', 'Budgeted Job Margin %',
              'Projected Gross Margin %', 'GP Variance(%)']:
        if c in proj_show_cols: proj_fmt[c] = '{:.1%}'

    st.dataframe(proj_display[proj_show_cols].style.format(proj_fmt), use_container_width=True, hide_index=True, height=min(500, 56+35*len(proj_display)))

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("💾  Save Projections to WIP.xlsx", use_container_width=True, key="save_proj"):
        try:
            save_df = st.session_state.master_wip.copy()
            save_cols = [
                'Current Month JTD', remaining_col, total_cost_proj_col,
                total_remaining_bill_col, billings_pct_col,
                '% Complete', jtd_revenue_col, 'JTD Billings',
                '(Over) / Under Billings', 'Over / (Under) Billings %',
                'Retainage', 'Net Billings',
                'Budgeted Job Margin %', 'Budgeted Gross Profit',
                'Projected Gross Margin %', 'Projected Gross Profit',
                'GP Variance(%)', 'GP Variance (USD)', 'Unearned Rev',
            ]
            for col in save_cols:
                if col in proj_display.columns:
                    col_map_dict = proj_display.set_index('Job #')[col].to_dict()
                    save_df[col] = save_df['Job #'].astype(str).str.strip().map(lambda j, m=col_map_dict: m.get(j, None))
            save_wip(save_df)
            st.session_state.master_wip = save_df.copy()
            st.success("Projections saved to " + WIP_FILE)
            wip_download_button("dl_projected")
        except PermissionError:
            st.error("Close WIP.xlsx before saving.")
        except Exception as e:
            st.error("Save failed: " + str(e))
